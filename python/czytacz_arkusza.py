# Program z zajęć
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = load_workbook('/Users/aleks/bootcamp/przestepstwa.xlsx')

def przerobarkusz(typprzestepstw):
    dane = {}
    ws = wb.get_sheet_by_name(typprzestepstw)
    #for row in ws.iter_rows('A{}:D{}'.format(2,ws.max_row)):

    for row in range(22, 345):
        komenda = ws.cell(column=1, row=row).value
        rok = ws.cell(column=2, row=row).value
        wykrycie = ws.cell(column=5, row=row).value

        if rok not in dane: dane[rok] = {}
        if komenda not in dane[rok]: dane[rok][komenda] = {}

        dane[rok][komenda] = wykrycie

if __name__ == "__main__":
    przerobarkusz('drogowe')
    przerobarkusz('uszkodzenia')
