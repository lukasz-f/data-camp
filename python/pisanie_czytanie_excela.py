# Program z zajęć
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = load_workbook('/Users/aleks/bootcamp/przestepstwa.xlsx')
dest_filename = '/Users/aleks/bootcamp/wyniki.xlsx'

def przerobarkusz(typprzestepstw):
    dane = {}
    ws = wb.get_sheet_by_name(typprzestepstw)
    #for row in ws.iter_rows('A{}:D{}'.format(2,ws.max_row)):

    for row in range(22, 346):
        komenda = ws.cell(column=1, row=row).value
        rok = ws.cell(column=2, row=row).value
        wykrycie = ws.cell(column=5, row=row).value

        if rok not in dane: dane[rok] = {}
        if komenda not in dane[rok]: dane[rok][komenda] = {}

        dane[rok][komenda] = wykrycie

    wbw = Workbook()
    wsw = wbw.create_sheet(title="wyniki")
    row=1
    for rok in dane:
        minimum_komenda = None
        minimum_wykrycie = 100
        for komenda,wykrycie in dane[rok].items():
            if wykrycie < minimum_wykrycie:
                minimum_wykrycie = wykrycie
                minimum_komenda = komenda

        print(rok, minimum_komenda, minimum_wykrycie)
        wsw.cell(column=1, row=row, value=rok)
        wsw.cell(column=2, row=row, value=minimum_komenda)
        wsw.cell(column=3, row=row, value=minimum_wykrycie)
        row += 1

    wbw.save(filename=dest_filename)

if __name__ == "__main__":
    nazwa = 'uszkodzenie'
    przerobarkusz(nazwa)
